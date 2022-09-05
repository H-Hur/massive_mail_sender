# -*- coding: cp949 -*- 
# Original file: https://underflow101.tistory.com/33
# ������� ����
# �Ʒ� ���̺귯������ ���̽� �⺻ ���� ���̺귯���̹Ƿ� ������ ��ġ�� �ʿ� �����ϴ�.

### Partly modified by Carrotday
### ���̹� ���Ͽ�

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
### ���� ������ �ѱ����Ͽ��� �о������ �Ʒ� ���̺귯�� ��ġ�� �ʿ��մϴ�. 
import olefile


SMTP_SERVER = "smtp.naver.com"
SMTP_PORT = 465
SMTP_USER = "Naver_ID@naver.com"
SMTP_PASSWORD = "NAVER_PW"
sender = SMTP_USER 

# smtp�� ������ ���� ������ ���� Ŭ�������� ����
smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
#smtp.ehlo()
if SMTP_PORT == 587: smtp.starttls()
# �ش� ������ �α���
smtp.login(SMTP_USER, SMTP_PASSWORD)

# ���� �Ʒ� ���� ��ȿ�� �˻� �Լ����� False�� ������ ������ ������ �ʽ��ϴ�.
def is_valid(addr):
    import re
    if re.match('(^[a-zA-Z-0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        return True
    else:
        return False
# �̸��� ������ �Լ�
def send_mail(sender,addr,subj,cont,attach1=None,attach2=None):
    if not is_valid(addr):
        print("Wrong email: " + addr)
        return
    
    # �ؽ�Ʈ ����
    msg = MIMEMultipart("alternative")
    # ÷�������� �ִ� ��� mixed�� multipart ����
    if attach1:
        msg = MIMEMultipart('mixed')
    msg["From"] = sender
    msg["To"] = addr
    msg["Subject"] = subj
    contents = cont
    text = MIMEText(_text = contents, _charset = "utf-8")
    msg.attach(text)

    if attach1:
        from email.mime.base import MIMEBase
        from email import encoders
        file_data = MIMEBase("application", "octect-stream")
        file_data.set_payload(open(attach1, "rb").read())
        encoders.encode_base64(file_data)
        import os
        filename = os.path.basename(attach1)
        file_data.add_header("Content-Disposition", 'attachment', filename=('UTF-8', '', filename))
        msg.attach(file_data)
    ### �� ��° ÷�������� �ִ� ���
    if attach2:
        file_data = MIMEBase("application", "octect-stream")
        file_data.set_payload(open(attach2, "rb").read())
        encoders.encode_base64(file_data)
        filename = os.path.basename(attach2)
        file_data.add_header("Content-Disposition", 'attachment', filename=('UTF-8', '', filename))
        msg.attach(file_data)    
    # ���� �߼�
    smtp.sendmail(sender, addr, msg.as_string())

# ���� ���Ͽ� ������ ������� �Ѳ����� ���� ��
# �Ʒ� openpyxl ���̺귯���� �ܺ� ���̺귯���̹Ƿ� pip3�� ���� ��ġ �� ����Ͻñ� �ٶ��ϴ�.
from openpyxl import load_workbook
#ws = load_workbook('send_test.xlsx').active
ws = load_workbook('send.xlsx').active

for row in ws.iter_rows():
    addr = row[0].value
    subj = row[1].value

### �׼��� ������ �״�� �������� �Ʒ� �� ���� ������.
#    cont = row[2].value
#    cont_file= row[2].value
### �׼��� �ѱ����� �̸��� �ְ� �ѱ����Ͽ� ���� ������ �������� �� �� �� ��� �Ʒ�  �� ���� ������
    file=olefile.OleFileIO( row[2].value)
    encoded_text = file.openstream('PrvText').read()
    cont = encoded_text.decode('UTF-16')

### �⺻ ÷�������� 1��(D��), �ִ� 2��(E��)
    attach1 = row[3].value
### �׼� E ���� �����̽��� ������ ���� �߻� 
    if  row[4].value != '': attach2 = row[4].value 
    if attach2 != '': send_mail(sender, addr, subj, cont, attach1, attach2) 
    else: send_mail(sender, addr, subj, cont, attach1)

# �ݱ�
smtp.close()

